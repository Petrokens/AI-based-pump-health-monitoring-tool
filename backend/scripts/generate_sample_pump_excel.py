"""
Generate sample Excel files for 2 pumps. Each pump has 3 files:
- pump_master.xlsx
- operation_log.xlsx
- maintenance_log.xlsx
"""
import os
from datetime import datetime, timedelta
from openpyxl import Workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
OUT_DIR = os.path.join(DATA_DIR, "sample_pumps")
os.makedirs(OUT_DIR, exist_ok=True)

# Pump 1: Centrifugal Process Pump / End Suction Overhung / API 610 OH1
PUMP1_MASTER = {
    "pump_id": "SAMPLE-01",
    "pump_type": "Centrifugal Process Pump",
    "pump_type_detail": "End Suction Overhung",
    "manufacturer": "Generic",
    "model": "ESO-200",
    "installation_date": "2023-01-15",
    "rated_power_kw": 45.0,
    "rated_flow_m3h": 350.0,
    "rated_head_m": 120.0,
    "flow_range_min_m3h": 100.0,
    "flow_range_max_m3h": 400.0,
    "head_range_min_m": 80.0,
    "head_range_max_m": 140.0,
    "rated_rpm": 2980,
    "seal_type": "Single_Mechanical",
    "bearing_type_de": "Roller",
    "bearing_type_nde": "Roller",
    "impeller_type": "Closed",
    "location": "Process_Unit_1",
    "criticality_level": "High",
    "serial_number": "ESO200-230115",
    "warranty_expiry": "2028-01-15",
    "last_overhaul": "2024-06-01",
    "min_safe_flow_m3h": 90.0,
    "max_safe_flow_m3h": 380.0,
    "max_motor_load_kw": 50.0,
    "max_suction_pressure_bar": 4.0,
    "efficiency_bep_percent": 84.0,
    "npshr_m": 3.5,
    "health_score": 88,
}

# Pump 2: Reciprocating Plunger Pump / Triplex Quintuplex / API 674
PUMP2_MASTER = {
    "pump_id": "SAMPLE-02",
    "pump_type": "Reciprocating Plunger Pump",
    "pump_type_detail": "Triplex / Quintuplex",
    "manufacturer": "Generic",
    "model": "TPL-150",
    "installation_date": "2023-03-20",
    "rated_power_kw": 75.0,
    "rated_flow_m3h": 25.0,
    "rated_head_m": 450.0,
    "flow_range_min_m3h": 5.0,
    "flow_range_max_m3h": 30.0,
    "head_range_min_m": 300.0,
    "head_range_max_m": 500.0,
    "rated_rpm": 350,
    "seal_type": "Packed_Plunger",
    "bearing_type_de": "Roller",
    "bearing_type_nde": "Roller",
    "impeller_type": "N/A",
    "location": "Process_Unit_2",
    "criticality_level": "High",
    "serial_number": "TPL150-230320",
    "warranty_expiry": "2028-03-20",
    "last_overhaul": "2024-09-10",
    "min_safe_flow_m3h": 5.0,
    "max_safe_flow_m3h": 28.0,
    "max_motor_load_kw": 82.0,
    "max_suction_pressure_bar": 2.0,
    "efficiency_bep_percent": 92.0,
    "npshr_m": 2.0,
    "health_score": 85,
}

MASTER_COLS = [
    "pump_id", "pump_type", "pump_type_detail", "manufacturer", "model", "installation_date",
    "rated_power_kw", "rated_flow_m3h", "rated_head_m", "flow_range_min_m3h", "flow_range_max_m3h",
    "head_range_min_m", "head_range_max_m", "rated_rpm", "seal_type", "bearing_type_de", "bearing_type_nde",
    "impeller_type", "location", "criticality_level", "serial_number", "warranty_expiry", "last_overhaul",
    "min_safe_flow_m3h", "max_safe_flow_m3h", "max_motor_load_kw", "max_suction_pressure_bar",
    "efficiency_bep_percent", "npshr_m", "health_score",
]

OP_LOG_COLS = [
    "timestamp", "pump_id", "flow_m3h", "discharge_pressure_bar", "suction_pressure_bar",
    "rpm", "motor_power_kw", "vibration_mm_s", "bearing_temp_c", "displacement_um", "status",
]

MNT_COLS = ["date", "pump_id", "action", "component", "notes", "downtime_hours"]


def write_master(path, row):
    wb = Workbook()
    ws = wb.active
    ws.title = "pump_master"
    ws.append(MASTER_COLS)
    ws.append([row.get(c) for c in MASTER_COLS])
    wb.save(path)


def write_operation_log(path, pump_id, num_rows=50):
    wb = Workbook()
    ws = wb.active
    ws.title = "operation_log"
    ws.append(OP_LOG_COLS)
    base = datetime(2025, 1, 1, 0, 0, 0)
    import random
    statuses = ["running", "running", "running", "standby", "alarm"]
    for i in range(num_rows):
        t = base + timedelta(minutes=15 * i)
        flow = round(180 + random.uniform(-40, 80), 2)
        dp = round(5 + random.uniform(0.5, 3), 2)
        sp = round(2 + random.uniform(0.5, 2), 2)
        rpm = round(2950 + random.uniform(-50, 50), 1)
        power = round(flow * 0.25 + random.uniform(-5, 5), 2)
        vib = round(1.5 + random.uniform(0, 1.5), 2)
        temp = round(55 + random.uniform(-5, 15), 1)
        disp = round(35 + random.uniform(0, 25), 1)
        status = random.choice(statuses)
        ws.append([t.isoformat(), pump_id, flow, dp, sp, rpm, power, vib, temp, disp, status])
    wb.save(path)


def write_maintenance_log(path, pump_id, num_rows=12):
    wb = Workbook()
    ws = wb.active
    ws.title = "maintenance_log"
    ws.append(MNT_COLS)
    base = datetime(2025, 1, 5)
    actions = ["inspect", "calibrate", "repair", "clean", "replace"]
    components = ["bearing", "seal", "alignment", "motor", "impeller"]
    notes_tpl = [
        "Routine inspection per PM schedule",
        "Alignment checked with laser kit",
        "Seal flush conductivity within spec",
        "Bearing housing vibration within limit",
        "Lubricant sample sent for analysis",
    ]
    for i in range(num_rows):
        d = (base + timedelta(days=14 * i)).strftime("%Y-%m-%d")
        action = actions[i % len(actions)]
        component = components[i % len(components)]
        notes = notes_tpl[i % len(notes_tpl)]
        downtime = (i % 3) * 2
        ws.append([d, pump_id, action, component, notes, downtime])
    wb.save(path)


def main():
    # Pump 1 (Centrifugal)
    p1_dir = os.path.join(OUT_DIR, "pump_1_centrifugal")
    os.makedirs(p1_dir, exist_ok=True)
    write_master(os.path.join(p1_dir, "pump_master.xlsx"), PUMP1_MASTER)
    write_operation_log(os.path.join(p1_dir, "operation_log.xlsx"), PUMP1_MASTER["pump_id"])
    write_maintenance_log(os.path.join(p1_dir, "maintenance_log.xlsx"), PUMP1_MASTER["pump_id"])

    # Pump 2 (Reciprocating)
    p2_dir = os.path.join(OUT_DIR, "pump_2_reciprocating")
    os.makedirs(p2_dir, exist_ok=True)
    write_master(os.path.join(p2_dir, "pump_master.xlsx"), PUMP2_MASTER)
    write_operation_log(os.path.join(p2_dir, "operation_log.xlsx"), PUMP2_MASTER["pump_id"])
    write_maintenance_log(os.path.join(p2_dir, "maintenance_log.xlsx"), PUMP2_MASTER["pump_id"])

    print("Created sample Excel files in:", OUT_DIR)
    print("  pump_1_centrifugal: pump_master.xlsx, operation_log.xlsx, maintenance_log.xlsx")
    print("  pump_2_reciprocating: pump_master.xlsx, operation_log.xlsx, maintenance_log.xlsx")


if __name__ == "__main__":
    main()
